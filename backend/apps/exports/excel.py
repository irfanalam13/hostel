from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def wb_from_rows(sheet_name: str, headers: list[str], rows: list[list]):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(headers)
    for r in rows:
        ws.append(r)

    # basic autosize
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(12, len(headers[col - 1]) + 2)

    return wb